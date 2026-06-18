#!/usr/bin/env python3
"""
BONK BLUEPRINT PROFITABILITY SCANNER  (v2, SDE-driven / unattended)
==================================================================
Answers ONE question: "What is the most profitable thing to MANUFACTURE, ranked by
ISK per hour?" so you know what to TRAIN toward and what to MINE.

WHAT CHANGED IN v2
  - Data layer rebuilt on EVE's SDE CSV dumps (Fuzzwork /dump/latest/csv/): blueprints,
    materials, build times, type names, and categories are read in BULK and cached
    locally, instead of one slow HTTP call per item. The old per-item endpoints had
    broken (category listing returned empty; build time vanished from the blueprint API),
    so v1 produced zero output. v2 is both correct and far faster.
  - Output QUANTITY is now honored: a blueprint run that makes 100 charges is valued at
    100x the unit price (v1 valued one unit, badly understating ammo/drone builds).
  - Names come from the SDE (no ESI). Prices still from Fuzzwork Jita market aggregates.
  - Unattended: no interactive menu, utf-8 console, real exit codes, a published HTML page
    (client-side encrypted, fail-closed), atomic writes. Monthly cadence.

HOW IT WORKS (all public data, no login):
  1. Build a blueprint index from the SDE (manufacturing activity only), for Ships,
     Modules, Ammo/Charges, Drones (published items).
  2. Pull live Jita prices for every product + input material.
  3. profit = product_sell * output_qty - material_cost(ME-adjusted) - rough job fee
  4. ISK/hour = profit / build_hours ; rank best-first.

USAGE:
    python bonk_blueprint_scanner.py --me 10
    python bonk_blueprint_scanner.py --categories ships,modules --top 100
    python bonk_blueprint_scanner.py --demo        # offline self-test

Pure stdlib + openpyxl (xlsx) + cryptography (page lock). Live runs need internet to
fuzzwork.co.uk. "Profit" assumes Jita fills; treat as a STRONG directional guide.
"""
import argparse, csv, io, json, math, os, sys, tempfile, time
import urllib.request, urllib.error
from datetime import datetime, timezone

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAVE_XLSX = True
except ImportError:
    HAVE_XLSX = False

UA = "BONK-BlueprintScanner/2.0 (Crown & Oak Capital; salesmaxxllc@gmail.com)"
SDE = "https://www.fuzzwork.co.uk/dump/latest/csv/"
MARKET = "https://market.fuzzwork.co.uk/aggregates/"
JITA_REGION = 10000002
CATEGORIES = {6: "Ships", 7: "Modules", 8: "Ammo", 18: "Drones"}
JOB_FEE_FACTOR = 0.05
SDE_CACHE = "sde_index.json"
OUT_XLSX = "bonk_blueprints.xlsx"
OUT_CSV = "bonk_blueprints.csv"
HTML = "index.html"


# ----------------------------------------------------------------------------
# HTTP
# ----------------------------------------------------------------------------

def fetch_text(url, timeout=90, retries=3):
    last = None
    for a in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read().decode("utf-8-sig", "replace")   # utf-8-sig strips the BOM
        except Exception as e:
            last = e
            time.sleep(2 * (a + 1))
    raise last


def fetch_json(url, timeout=60, retries=3):
    return json.loads(fetch_text(url, timeout, retries))


def _rows(text):
    return csv.DictReader(io.StringIO(text))


# ----------------------------------------------------------------------------
# SDE BLUEPRINT INDEX (bulk, cached)
# ----------------------------------------------------------------------------

def build_index(refresh=False, max_cache_days=25):
    """Return {product_type_id: {bp_id, name, category, build_seconds, out_qty, materials}}.
    Built from the SDE CSV dumps and cached locally."""
    if not refresh and os.path.exists(SDE_CACHE):
        try:
            obj = json.load(open(SDE_CACHE, encoding="utf-8"))
            built = datetime.fromisoformat(obj["built"])
            if (datetime.now(timezone.utc) - built).days < max_cache_days:
                prods = {int(k): v for k, v in obj["products"].items()}
                nms = {int(k): v for k, v in obj.get("names", {}).items()}
                return prods, nms
        except Exception:
            pass

    print("  Downloading EVE SDE from Fuzzwork (bulk, cached for "
          f"{max_cache_days}d)...")
    # group -> category (published groups)
    groups = {}
    for row in _rows(fetch_text(SDE + "invGroups.csv")):
        if row.get("published") == "1":
            try:
                groups[int(row["groupID"])] = int(row["categoryID"])
            except (ValueError, TypeError):
                pass
    # published types in our target categories -> (name, category_id);
    # allnames -> every typeID's name (needed to label build materials, which are
    # minerals/components that live outside our target categories)
    types = {}
    allnames = {}
    for row in _rows(fetch_text(SDE + "invTypes.csv")):
        try:
            tid = int(row["typeID"])
        except (ValueError, TypeError):
            continue
        allnames[tid] = row.get("typeName", f"id:{tid}")
        if row.get("published") != "1":
            continue
        try:
            gid = int(row["groupID"])
        except (ValueError, TypeError):
            continue
        cat = groups.get(gid)
        if cat in CATEGORIES:
            types[tid] = (allnames[tid], cat)
    # blueprint -> (product, out_qty) for manufacturing (activity 1)
    bp_product = {}
    for row in _rows(fetch_text(SDE + "industryActivityProducts.csv")):
        if row.get("activityID") != "1":
            continue
        try:
            bp_product[int(row["typeID"])] = (int(row["productTypeID"]), int(row["quantity"]))
        except (ValueError, TypeError):
            pass
    # blueprint -> build seconds (activity 1)
    bp_time = {}
    for row in _rows(fetch_text(SDE + "industryActivity.csv")):
        if row.get("activityID") != "1":
            continue
        try:
            bp_time[int(row["typeID"])] = int(row["time"])
        except (ValueError, TypeError):
            pass
    # blueprint -> [(material, qty)] (activity 1)
    bp_mats = {}
    for row in _rows(fetch_text(SDE + "industryActivityMaterials.csv")):
        if row.get("activityID") != "1":
            continue
        try:
            bp_mats.setdefault(int(row["typeID"]), []).append(
                (int(row["materialTypeID"]), int(row["quantity"])))
        except (ValueError, TypeError):
            pass
    # meta group, to keep TECH I only. Faction/officer/deadspace/special-edition items
    # (meta 3/4/5/6/...) have SDE blueprints but trade as rare collectibles, producing
    # absurd fake margins. Tech I = metaGroupID 1, or absent (base T1 items).
    metatypes = {}
    for row in _rows(fetch_text(SDE + "invMetaTypes.csv")):
        try:
            metatypes[int(row["typeID"])] = int(row["metaGroupID"])
        except (ValueError, TypeError):
            pass

    products = {}
    for bp_id, (prod_id, out_qty) in bp_product.items():
        if prod_id not in types:
            continue
        if metatypes.get(prod_id, 1) != 1:        # Tech I only
            continue
        secs, mats = bp_time.get(bp_id), bp_mats.get(bp_id)
        if not secs or not mats:
            continue
        name, cat = types[prod_id]
        products[prod_id] = {"bp_id": bp_id, "name": name, "category": CATEGORIES[cat],
                             "build_seconds": secs, "out_qty": out_qty, "materials": mats}

    # name lookup for the products + their materials (small subset, not all 110k types)
    needed = set(products)
    for p in products.values():
        needed.update(m for m, _ in p["materials"])
    names = {i: allnames.get(i, f"id:{i}") for i in needed}

    try:
        with open(SDE_CACHE, "w", encoding="utf-8") as f:
            json.dump({"built": datetime.now(timezone.utc).isoformat(),
                       "products": {str(k): v for k, v in products.items()},
                       "names": {str(k): v for k, v in names.items()}}, f)
    except Exception:
        pass
    print(f"  Indexed {len(products)} manufacturable items.")
    return products, names


# ----------------------------------------------------------------------------
# PRICES (Fuzzwork Jita market aggregates)
# ----------------------------------------------------------------------------

def get_prices(type_ids):
    prices, ids = {}, sorted({int(t) for t in type_ids if t})
    for i in range(0, len(ids), 100):
        chunk = ids[i:i + 100]
        url = f"{MARKET}?region={JITA_REGION}&types=" + ",".join(map(str, chunk))
        try:
            data = fetch_json(url)
        except Exception:
            continue
        if not isinstance(data, dict):
            continue
        for tid, info in data.items():
            try:
                sell = float((info.get("sell") or {}).get("min") or 0)
                buy = float((info.get("buy") or {}).get("max") or 0)
                vol = float((info.get("sell") or {}).get("volume") or 0)
                prices[int(tid)] = {"sell": sell, "buy": buy, "volume": vol}
            except (ValueError, TypeError):
                pass
        time.sleep(0.1)
    return prices


def compute_profit(pid, product, prices, me_factor):
    pp = prices.get(pid)
    if not pp or pp["sell"] <= 0:
        return None
    secs = product["build_seconds"]
    if not secs or secs <= 0:
        return None
    mat_cost = 0.0
    for mid, qty in product["materials"]:
        mp = prices.get(mid)
        if not mp or mp["sell"] <= 0:
            return None                       # an unpriceable input -> skip the build
        mat_cost += qty * me_factor * mp["sell"]
    if mat_cost <= 0:
        return None
    out_qty = product.get("out_qty", 1) or 1
    product_value = pp["sell"] * out_qty       # honor multi-unit runs (e.g. 100 charges)
    job_fee = mat_cost * JOB_FEE_FACTOR
    profit = product_value - mat_cost - job_fee
    hours = secs / 3600.0
    return {
        "type_id": pid, "name": product["name"], "category": product["category"],
        "out_qty": out_qty, "unit_sell": pp["sell"], "unit_mat": mat_cost / out_qty,
        "product_value": product_value, "material_cost": mat_cost, "profit": profit,
        "build_hours": hours, "isk_per_hour": (profit / hours if hours > 0 else 0),
        "margin_pct": (profit / (mat_cost + job_fee) * 100) if (mat_cost + job_fee) > 0 else 0,
        "daily_volume": pp["volume"],
    }


def mat_summary(product, me_factor, names, limit=10):
    """ME-adjusted material requirements for one build run, labeled and sorted by
    quantity (the big mineral inputs first). Used for the 'needs:' line on the page."""
    out = []
    for mid, qty in product["materials"]:
        adj = max(1, int(round(qty * me_factor)))
        out.append({"name": names.get(mid, f"id:{mid}"), "qty": adj})
    out.sort(key=lambda m: m["qty"], reverse=True)
    return out[:limit]


# mineral -> (best accessible ore, security tier, refined yield per ore unit, ore m3/unit).
# "Accessibility-aware": the richest ore in the LOWEST security tier where the mineral is
# reasonably mined (so a highsec corp gets Veldspar for Tritanium, not nullsec Spodumain).
# Yields are from the SDE reprocessing tables; minerals only found in low/null are flagged.
MINERAL_ORE = {
    "Tritanium": ("Veldspar",    "Hi",   4.0,  0.10),
    "Pyerite":   ("Scordite",    "Hi",   1.1,  0.15),
    "Mexallon":  ("Plagioclase", "Hi",   0.7,  0.35),
    "Isogen":    ("Kernite",     "Hi",   1.2,  1.20),
    "Nocxium":   ("Crokite",     "Lo",   8.0,  16.0),
    "Zydrine":   ("Bistot",      "Null", 1.6,  16.0),
    "Megacyte":  ("Arkonor",     "Null", 1.2,  16.0),
    "Morphite":  ("Mercoxit",    "Null", 1.4,  40.0),
}


def build_path(product, me_factor, names):
    """A clear 'go mine this' plan: split the ME-adjusted bill of materials into minerals
    (mineable) and components (build/buy), and for each mineral pick the best ore + how many
    units to mine. Per-mineral quantities are an upper bound (ores drop byproducts)."""
    minerals, components = [], []
    for mid, qty in product["materials"]:
        nm = names.get(mid, f"id:{mid}")
        adj = max(1, int(round(qty * me_factor)))
        (minerals if nm in MINERAL_ORE else components).append((nm, adj))
    minerals.sort(key=lambda x: x[1], reverse=True)
    components.sort(key=lambda x: x[1], reverse=True)
    plan, total_m3 = [], 0.0
    for nm, qty in minerals:
        ore, sec, ypu, vol = MINERAL_ORE[nm]
        units = int(math.ceil(qty / ypu)) if ypu else 0
        m3 = units * vol
        total_m3 += m3
        plan.append({"mineral": nm, "qty": qty, "ore": ore, "sec": sec, "units": units, "m3": m3})
    return {"plan": plan, "components": components, "total_m3": total_m3}


# ----------------------------------------------------------------------------
# OUTPUT
# ----------------------------------------------------------------------------

def _atomic(path, write_fn, suffix):
    d = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp = tempfile.mkstemp(suffix=suffix, dir=d)
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as f:
            write_fn(f)
        os.replace(tmp, path)
    except Exception:
        try: os.remove(tmp)
        except OSError: pass
        raise
    return path


CSV_COLS = ["rank", "name", "category", "isk_per_hour", "profit", "build_hours",
            "margin_pct", "product_value", "material_cost", "daily_volume", "type_id"]


def write_csv(path, rows):
    def w(f):
        wr = csv.writer(f)
        wr.writerow(CSV_COLS)
        for r in rows:
            wr.writerow([r.get("rank"), r.get("name"), r.get("category"),
                         round(r["isk_per_hour"]), round(r["profit"]), round(r["build_hours"], 2),
                         round(r["margin_pct"], 1), round(r["product_value"]),
                         round(r["material_cost"]), int(r["daily_volume"]), r.get("type_id")])
    _atomic(path, w, ".csv")


def write_xlsx(path, rows, me_level):
    if not HAVE_XLSX:
        print("  (openpyxl not installed; skipping .xlsx)")
        return None
    wb = Workbook(); ws = wb.active; ws.title = f"Best Builds ME{me_level}"
    hdr = ["Rank", "Item", "Category", "ISK/Hour", "Profit/Build", "Build Hrs",
           "Margin %", "Product Sell", "Mat Cost", "Vol/Day", "Type ID"]
    ws.append(hdr)
    fill = PatternFill("solid", fgColor="0A110D")
    font = Font(bold=True, color="5FD9A0", name="Arial", size=11)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=1, column=c); cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get("rank"), r["name"], r["category"], round(r["isk_per_hour"]),
                   round(r["profit"]), round(r["build_hours"], 2), round(r["margin_pct"], 1),
                   round(r["product_value"]), round(r["material_cost"]),
                   int(r["daily_volume"]), r["type_id"]])
    for col in (4, 5, 8, 9, 10):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "#,##0"
    for i, wdt in enumerate([6, 32, 10, 16, 16, 10, 9, 14, 14, 10, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{ws.max_row}"
    d = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=d); os.close(fd)
    try:
        wb.save(tmp); os.replace(tmp, path); return path
    except PermissionError:
        try: os.remove(tmp)
        except OSError: pass
        print(f"  ({os.path.basename(path)} open in Excel? skipped.)"); return None
    except Exception as e:
        try: os.remove(tmp)
        except OSError: pass
        print(f"  (xlsx failed: {e})"); return None


# ----------------------------------------------------------------------------
# DEMO
# ----------------------------------------------------------------------------

def run_demo(args):
    import random
    random.seed(17)
    cats = list(CATEGORIES.values())
    rows = []
    for i in range(1, 61):
        prof = random.choice([2e5, 1e6, 5e6, 2e7]) * random.uniform(0.4, 1.8)
        hrs = random.uniform(0.1, 6)
        mat = random.uniform(1e6, 5e8)
        rows.append({"type_id": 1000 + i, "name": f"Demo Item {i}", "category": random.choice(cats),
                     "product_value": mat + prof, "material_cost": mat, "profit": prof,
                     "build_hours": hrs, "isk_per_hour": prof / hrs,
                     "margin_pct": prof / mat * 100, "daily_volume": random.randint(0, 5000)})
    rows = _top_per_category(rows, args.top)
    _finalize(rows, args, me_level=args.me if args.me is not None else 10, demo=True)
    print(f"\n  DEMO: {len(rows)} fake builds (no network). Outputs written with _demo suffix.\n")
    return 0


# ----------------------------------------------------------------------------
# FINALIZE
# ----------------------------------------------------------------------------

CAT_ORDER = ["Ships", "Modules", "Ammo", "Drones"]


def _top_per_category(results, n):
    """Group by category, keep the top n by ISK/hour in each (rank resets per
    category). Returned rows are ordered Ships, Modules, Ammo, Drones."""
    by = {}
    for r in results:
        by.setdefault(r["category"], []).append(r)
    out, cats = [], [c for c in CAT_ORDER if c in by] + [c for c in by if c not in CAT_ORDER]
    for cat in cats:
        grp = sorted(by[cat], key=lambda x: x["isk_per_hour"], reverse=True)[:n]
        for i, r in enumerate(grp, 1):
            r["rank"] = i
        out.extend(grp)
    return out


def _finalize(rows, args, me_level, demo=False):
    basis = "buy materials at Jita sell, sell product at Jita sell, ~5% job fee"
    suffix = "_demo" if demo else ""
    out_csv = OUT_CSV.rsplit(".", 1)[0] + suffix + ".csv"
    out_xlsx = OUT_XLSX.rsplit(".", 1)[0] + suffix + ".xlsx"
    if not args.no_xlsx and write_xlsx(out_xlsx, rows, me_level):
        print(f"  Wrote workbook {out_xlsx}")
    write_csv(out_csv, rows)
    print(f"  Wrote {len(rows)} rows to {out_csv}")

    if not args.no_html:
        import blueprint_page
        html_file = "index_demo.html" if demo else args.html
        pw = os.environ.get("EVE_PAGE_PASSWORD")
        state = {"generated_at": datetime.now(timezone.utc).isoformat(), "me_level": me_level,
                 "basis": basis, "count": len(rows), "per_cat": args.top, "rows": rows}
        blueprint_page.write_html(html_file, state, pw, allow_plain=(demo or args.allow_unlocked))
        lock = "locked" if pw else ("UNLOCKED (allow-unlocked)" if (demo or args.allow_unlocked) else "?")
        print(f"  Wrote page {html_file}  [{lock}]")

    print(f"\n  Top {args.top} by ISK/hour per category (ME{me_level}):")
    cur = None
    for r in rows:
        if r.get("category") != cur:
            cur = r.get("category")
            print(f"\n  {str(cur).upper()}")
            print("  " + "-" * 68)
        ih = r["isk_per_hour"]
        s = f"{ih/1e6:.1f}M" if ih >= 1e6 else f"{ih/1e3:.0f}k"
        print(f"   {r.get('rank', 0):>2} {r['name'][:32]:<32} {s:>8}/hr  ({r['margin_pct']:.0f}% margin)")


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Rank the most profitable items to manufacture (ISK/hour).")
    ap.add_argument("--me", type=int, default=None, help="Material Efficiency level 0-10 (default 10)")
    ap.add_argument("--top", type=int, default=5, help="how many top builds to keep PER category")
    ap.add_argument("--page-top", type=int, default=5, help="(unused; page shows all per-category rows)")
    ap.add_argument("--categories", default="all", help="comma list: ships,modules,ammo,drones (default all)")
    ap.add_argument("--min-volume", type=float, default=20, help="drop illiquid products (sell volume below this); raise for bulk-only")
    ap.add_argument("--max-margin", type=float, default=1000.0,
                    help="drop builds above this %% margin (collector/data artifacts; default 1000)")
    ap.add_argument("--refresh", action="store_true", help="force re-download of the SDE index")
    ap.add_argument("--max-cache-days", type=int, default=25, help="rebuild SDE index if older than this")
    ap.add_argument("--out", default=OUT_XLSX)
    ap.add_argument("--html", default=HTML)
    ap.add_argument("--no-xlsx", action="store_true")
    ap.add_argument("--no-html", action="store_true")
    ap.add_argument("--allow-unlocked", action="store_true", help="permit an UNLOCKED page (local only)")
    ap.add_argument("--demo", action="store_true")
    args = ap.parse_args()

    me_level = max(0, min(10, args.me if args.me is not None else 10))
    me_factor = 1.0 - (me_level / 100.0)
    print(f"\n  BONK BLUEPRINT SCANNER v2  [{'DEMO' if args.demo else 'LIVE'}]  ME{me_level}")

    # fail-closed: never write an unlocked page on the publish path by accident
    pw = os.environ.get("EVE_PAGE_PASSWORD")
    if not args.demo and not args.no_html and not args.allow_unlocked and not (pw or "").strip():
        print("  ERROR: EVE_PAGE_PASSWORD is unset/blank; refusing to write an unlocked page.")
        print("  Set it, or pass --allow-unlocked (local only) / --no-html.")
        return 2

    if args.demo:
        return run_demo(args)

    # category filter
    if args.categories.lower() == "all":
        wanted = set(CATEGORIES.values())
    else:
        wanted = {c.strip().capitalize() for c in args.categories.split(",")}

    try:
        products, names = build_index(refresh=args.refresh, max_cache_days=args.max_cache_days)
    except Exception as e:
        print(f"  Could not build SDE index: {e}")
        return 1
    products = {pid: p for pid, p in products.items() if p["category"] in wanted}
    if not products:
        print("  No manufacturable items for the chosen categories.")
        return 1
    print(f"  {len(products)} candidate items in {sorted(wanted)}. Pulling Jita prices...")

    all_ids = set(products)
    for p in products.values():
        all_ids.update(m for m, _ in p["materials"])
        all_ids.add(p["bp_id"])   # also price the blueprint, to test it is obtainable
    prices = get_prices(all_ids)
    if not prices:
        print("  No price data returned (market endpoint down?).")
        return 1

    results = []
    for pid, p in products.items():
        # Skip items whose blueprint is not sold on the market: EDENCOM, Triglavian,
        # and other special items have SDE manufacturing rows but no obtainable BPO,
        # so a normal player cannot actually build them.
        if prices.get(p["bp_id"], {}).get("sell", 0) <= 0:
            continue
        calc = compute_profit(pid, p, prices, me_factor)
        if (calc and calc["isk_per_hour"] > 0
                and calc["daily_volume"] >= args.min_volume
                and calc["margin_pct"] <= args.max_margin):   # cap collector/artifact margins
            calc["materials"] = mat_summary(p, me_factor, names)
            calc["build_path"] = build_path(p, me_factor, names)
            results.append(calc)
    if not results:
        print("  No profitable items computed (sparse prices or all filtered).")
        return 1

    rows = _top_per_category(results, args.top)

    _finalize(rows, args, me_level)
    print(f"\n  {len(results)} profitable builds found; showing top {args.top} per category. "
          "Top of each list = what to train toward and mine for.\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
